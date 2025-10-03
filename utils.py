import openpyxl, re, os, numpy as np, matplotlib.pyplot as plt
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Side
from matplotlib.backends.backend_pdf import PdfPages 
from matplotlib.ticker import PercentFormatter
from datetime import datetime

# Importación opcional de IA
try:
    from informe_ia import generar_informe_ia, exportar_informe_pdf
    IA_DISPONIBLE = True
except ImportError:
    IA_DISPONIBLE = False
    print("⚠️ Módulo informe_ia no disponible. Se generará PDF sin análisis de IA.")

def find_year_in_workbook(wb):
    pattern = re.compile(r"\b(20\d{2})\b")
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=40, min_col=1, max_col=12):
            for cell in row:
                if cell.value and (m := pattern.search(str(cell.value))):
                    return int(m.group(1))
    raise ValueError("No se encontró año en el libro. Asegúrate que el archivo contiene el año (YYYY).")

def find_company_name(wb):
    """Extrae el nombre de la empresa del archivo Excel"""
    name = str(wb.active['A6'].value or "NOMBRE_DE_LA_EMPRESA").strip()
    return name.split(':', 1)[1].strip() if name.upper().startswith("EMPRESA:") else name

def read_range_values(wb, sheet_name, start_row, end_row, col='C'):
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    return [ws.cell(row=r, column=column_index_from_string(col)).value or 0 for r in range(start_row, end_row+1)]

def write_column_into_model(ws, target_start_row, values, col_idx):
    for i, val in enumerate(values):
        r = target_start_row + i
        cell = ws.cell(row=r, column=col_idx)
        cell.value = val
        if r != 3 and isinstance(val, (int, float)):
            cell.number_format = "#,##0"

def safe_output_path(output_dir, empresa, years_sorted):
    """Genera un nombre de archivo seguro y único"""
    empresa_limpia = re.sub(r'[<>:"/\\|?*]', '', empresa)[:50]
    base = f"REPORTE_{empresa_limpia}_{years_sorted[-1]}-{years_sorted[0]}.xlsx"
    out_file = os.path.join(output_dir, base)
    if os.path.exists(out_file):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_file = os.path.join(output_dir, f"REPORTE_{empresa_limpia}_{years_sorted[-1]}-{years_sorted[0]}_{ts}.xlsx")
    return out_file

def process_files_and_generate_report(input_paths, model_path="BASE.xlsx", output_dir="."):
    model_wb = openpyxl.load_workbook(model_path)
    
    # Obtener nombre de la empresa del primer archivo
    first_wb = openpyxl.load_workbook(input_paths[0], data_only=False)
    company_name = find_company_name(first_wb)
    first_wb.close()

    for style_name, fmt in [("percentage_style", "0.00%"), ("decimal_style", "0.00")]:
        try: model_wb.add_named_style(NamedStyle(name=style_name, number_format=fmt))
        except: pass

    entries, years = [], []
    for p in input_paths:
        wb = openpyxl.load_workbook(p, data_only=True)
        year = int(str(wb.active["C12"].value or find_year_in_workbook(wb)).strip())
        years.append(year)
        entries.append({'year': year, 'bs': read_range_values(wb, wb.sheetnames[0], 12, 88, 'C'),
                       'is': read_range_values(wb, wb.sheetnames[0], 91, 131, 'C'),
                       'cf': read_range_values(wb, wb.sheetnames[0], 185, 259, 'C')})
        wb.close()

    years_sorted = sorted(years, reverse=True)
    if len(set(years)) != len(years): raise ValueError("Años repetidos detectados.")
    for i in range(len(years_sorted)-1):
        if years_sorted[i] - years_sorted[i+1] != 1:
            raise ValueError(f"Los años deben ser consecutivos. Detectados: {', '.join(map(str, years_sorted))}")

    base_col = 3
    sheets = {'ESTADO DE SITUACIÓN FINANCIERA': 'bs', 'ESTADO DE RESULTADOS': 'is', 'ESTADO DE FLUJO DE EFECTIVO': 'cf'}
    for idx, e in enumerate(sorted(entries, key=lambda e: e['year'], reverse=True)):
        for sheet_name, key in sheets.items():
            write_column_into_model(model_wb[sheet_name], 3, e[key], base_col + idx)

            ws = model_wb[sheet_name]
            cell = ws.cell(row=3, column=base_col + idx, value=str(e['year']))
            cell.fill = PatternFill(start_color="FF337AB6", end_color="FF337AB6", fill_type="solid")
            cell.font = Font(color="FFFFFFFF", bold=True)

            thick_black_border = Border(
                left=Side(style='thin', color='FF000000'),
                right=Side(style='thin', color='FF000000'),
                top=Side(style='thin', color='FF000000'),
                bottom=Side(style='thin', color='FF000000')
            )
            for r in range(3, 3 + len(e[key])):
                ws.cell(row=r, column=base_col + idx).border = thick_black_border

    def vertical_formula(sheet, row, col, sheet_name):
        coord = sheet.cell(row=row, column=col).coordinate
        base_rows = {'ESTADO DE SITUACIÓN FINANCIERA': 40, 'ESTADO DE RESULTADOS': 4}
        if sheet_name in base_rows:
            return f"=IFERROR({coord}/{sheet.cell(row=base_rows[sheet_name], column=col).coordinate},0)" if row != base_rows[sheet_name] else None
        if sheet_name == 'ESTADO DE FLUJO DE EFECTIVO':
            base_row = 24 if row <= 24 else (52 if row <= 52 else (72 if row <= 72 else None))
            return f"=IFERROR({coord}/{sheet.cell(row=base_row, column=col).coordinate},0)" if base_row and row != base_row else None

    n_entries = len(entries)
    for sheet_name, start_row, rows_count in [('ESTADO DE SITUACIÓN FINANCIERA', 3, 77), 
                                                ('ESTADO DE RESULTADOS', 3, 41), 
                                                ('ESTADO DE FLUJO DE EFECTIVO', 3, 75)]:
        ws = model_wb[sheet_name]
        vert_start, hor_start = base_col + n_entries + 2, base_col + 2 * n_entries + 4
        title_fill = PatternFill(start_color="FF337AB6", end_color="FF337AB6", fill_type="solid")
        title_font = Font(color="FFFFFFFF", bold=True)
        ws.cell(row=2, column=vert_start, value="ANÁLISIS VERTICAL").fill = title_fill
        ws.cell(row=2, column=vert_start).font = title_font
        ws.cell(row=2, column=hor_start, value="ANÁLISIS HORIZONTAL").fill = title_fill
        ws.cell(row=2, column=hor_start).font = title_font

        header_fill = PatternFill(start_color="FF337AB6", end_color="FF337AB6", fill_type="solid")
        header_font = Font(color="FFFFFFFF", bold=True)
        for i, y in enumerate(years_sorted):
            cell = ws.cell(row=3, column=vert_start+i, value=str(y))
            cell.number_format = 'General'
            cell.fill = header_fill
            cell.font = header_font
        # Corregir: el análisis horizontal debe mostrar los años correctos (más reciente primero)
        for i in range(len(years_sorted)-1):
            cell = ws.cell(row=3, column=hor_start+i, value=str(years_sorted[i]))
            cell.number_format = 'General'
            cell.fill = header_fill
            cell.font = header_font
        
        for r in range(start_row+1, start_row+rows_count):
            for j in range(n_entries):
                col_idx = base_col + j
                if formula := vertical_formula(ws, r, col_idx, sheet_name):
                    ws.cell(row=r, column=vert_start+j, value=formula).style = "percentage_style"
                if j+1 < n_entries:
                    ws.cell(row=r, column=hor_start+j, 
                           value=f"=IFERROR(({ws.cell(row=r, column=col_idx).coordinate}-{ws.cell(row=r, column=base_col+j+1).coordinate})/{ws.cell(row=r, column=base_col+j+1).coordinate},0)"
                           ).style = "percentage_style"

    ws_ratios = model_wb['RATIOS']
    for merged in list(ws_ratios.merged_cells.ranges): ws_ratios.unmerge_cells(str(merged))
    
    header_fill = PatternFill(start_color="FF337AB6", end_color="FF337AB6", fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)
    for i, y in enumerate(years_sorted):
        cell = ws_ratios.cell(row=3, column=5+i, value=y)
        cell.number_format = '0'
        cell.fill = header_fill
        cell.font = header_font 
    
    for idx in range(len(years_sorted)):
        col, dc = 5 + idx, get_column_letter(base_col + idx)
        pc = get_column_letter(base_col + idx + 1) if idx < len(years_sorted) - 1 else dc
        
        for row, formula, style in [
            (4, f"=IFERROR('ESTADO DE SITUACIÓN FINANCIERA'!{dc}20/'ESTADO DE SITUACIÓN FINANCIERA'!{dc}55,0)", "decimal_style"),
            (5, f"=IFERROR(('ESTADO DE SITUACIÓN FINANCIERA'!{dc}20-'ESTADO DE SITUACIÓN FINANCIERA'!{dc}13)/'ESTADO DE SITUACIÓN FINANCIERA'!{dc}55,0)", "decimal_style"),
            (6, f"=IFERROR('ESTADO DE SITUACIÓN FINANCIERA'!{dc}69/'ESTADO DE SITUACIÓN FINANCIERA'!{dc}40,0)", "decimal_style"),
            (7, f"=IFERROR('ESTADO DE SITUACIÓN FINANCIERA'!{dc}69/'ESTADO DE SITUACIÓN FINANCIERA'!{dc}78,0)", "decimal_style"),
            (8, f"=IFERROR('ESTADO DE RESULTADOS'!{dc}28/'ESTADO DE RESULTADOS'!{dc}4,0)", "percentage_style"),
            (9, f"=IFERROR('ESTADO DE RESULTADOS'!{dc}28/'ESTADO DE SITUACIÓN FINANCIERA'!{dc}40,0)", "percentage_style"),
            (10, f"=IFERROR('ESTADO DE RESULTADOS'!{dc}28/'ESTADO DE SITUACIÓN FINANCIERA'!{dc}78,0)", "percentage_style"),
            (11, f"=IFERROR('ESTADO DE RESULTADOS'!{dc}4/'ESTADO DE SITUACIÓN FINANCIERA'!{dc}40,0)", "decimal_style")
        ]:
            ws_ratios.cell(row=row, column=col, value=formula).style = style
        
        if idx < len(years_sorted) - 1:
            ws_ratios.cell(row=12, column=col, value=f"=IFERROR('ESTADO DE RESULTADOS'!{dc}4/((SUM('ESTADO DE SITUACIÓN FINANCIERA'!{dc}8:{dc}11)+SUM('ESTADO DE SITUACIÓN FINANCIERA'!{dc}24:{dc}27)+SUM('ESTADO DE SITUACIÓN FINANCIERA'!{pc}8:{pc}11)+SUM('ESTADO DE SITUACIÓN FINANCIERA'!{pc}24:{pc}27))/2),0)").style = "decimal_style"
            ws_ratios.cell(row=13, column=col, value=f"=IFERROR(-'ESTADO DE RESULTADOS'!{dc}5/((('ESTADO DE SITUACIÓN FINANCIERA'!{dc}13+'ESTADO DE SITUACIÓN FINANCIERA'!{pc}13)+('ESTADO DE SITUACIÓN FINANCIERA'!{dc}30+'ESTADO DE SITUACIÓN FINANCIERA'!{pc}30))/2),0)").style = "decimal_style"
        else:
            for row in [12, 13]: ws_ratios.cell(row=row, column=col, value=0).style = "decimal_style"

    out_file = safe_output_path(output_dir, company_name, years_sorted)
    model_wb.save(out_file)
    model_wb.close()
    return out_file

def safe_float(x):
    try: return float(x) if x is not None else 0.0
    except: return 0.0

def add_value_labels(ax, rects, values, ttype):
    """Añade etiquetas de valores optimizadas"""
    y_min, y_max = ax.get_ylim()
    
    for rect, v in zip(rects, values):
        bar_height = rect.get_height()
        
        if bar_height >= 0:
            label_y = bar_height + (y_max - y_min) * 0.015
            va = 'bottom'
        else:
            label_y = bar_height - (y_max - y_min) * 0.015
            va = 'top'
        
        if ttype == 'pct':
            label_text = f"{v*100:.1f}%"
        else:
            label_text = f"{v:.2f}" if abs(v) < 10 else f"{v:.1f}"
        
        ax.text(rect.get_x() + rect.get_width()/2, label_y, label_text,
                ha='center', va=va, fontsize=8, fontweight='medium', 
                color='darkslategray')

def nice_ticks(ax, values, ttype, n_ticks=5):
    """Optimiza la escala de los ejes"""
    v_max, v_min = max(values + [0]), min(values + [0])
    
    if (span := v_max - v_min) == 0:
        v_max, v_min = v_max + abs(v_max)*0.1 + 0.1, v_min - abs(v_min)*0.1 - 0.1
        span = v_max - v_min
    
    margin_top = span * 0.2
    margin_bottom = span * 0.05
    
    ax.set_ylim(v_min - margin_bottom, v_max + margin_top)
    
    ticks = np.linspace(v_min - margin_bottom, v_max + margin_top, n_ticks)
    ax.set_yticks(ticks)
    
    if ttype == 'pct':
        ax.yaxis.set_major_formatter(PercentFormatter(1.0, decimals=1))
    else:
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, pos: f"{x:.1f}" if abs(x) < 10 else f"{x:.0f}"))

def generate_ratios_charts_pdf(report_path, output_dir):
    plt.rcParams.update({'font.sans-serif': 'Arial', 'font.size': 9})
    
    wb = openpyxl.load_workbook(report_path, data_only=True)
    ws_ratios, ws_bs, ws_is = wb['RATIOS'], wb['ESTADO DE SITUACIÓN FINANCIERA'], wb['ESTADO DE RESULTADOS']
    
    # Extraer nombre del archivo del reporte
    import re
    filename = os.path.basename(report_path)
    # Extraer nombre entre REPORTE_ y el año
    match = re.search(r'REPORTE_(.+?)_\d{4}', filename)
    company_name = match.group(1) if match else find_company_name(wb)

    years, col = [], 5
    while (v := ws_ratios.cell(row=3, column=col).value):
        years.append(int(v))
        col += 1
    if not years:
        wb.close()
        raise ValueError("No se encontraron años en la hoja RATIOS.")

    ratios_info = [
        {'row': 4, 'name': 'Liquidez corriente', 'short': 'Liquidez corriente', 'type': 'ratio', 'category': 'Liquidez'},
        {'row': 5, 'name': 'Prueba ácida', 'short': 'Prueba ácida', 'type': 'ratio', 'category': 'Liquidez'},
        {'row': 6, 'name': 'Razón de deuda total', 'short': 'Razón deuda total', 'type': 'ratio', 'category': 'Endeudamiento'},
        {'row': 7, 'name': 'Razón deuda/patrimonio', 'short': 'Razón deuda/patrimonio', 'type': 'ratio', 'category': 'Endeudamiento'},
        {'row': 8, 'name': 'Margen neto', 'short': 'Margen neto', 'type': 'pct', 'category': 'Rentabilidad'},
        {'row': 9, 'name': 'ROA', 'short': 'ROA', 'type': 'pct', 'category': 'Rentabilidad'},
        {'row': 10, 'name': 'ROE', 'short': 'ROE', 'type': 'pct', 'category': 'Rentabilidad'},
        {'row': 11, 'name': 'Rotación de activos totales', 'short': 'Rotación activos', 'type': 'ratio', 'category': 'Actividad'},
        {'row': 12, 'name': 'Rotación de cuentas por cobrar', 'short': 'Rotación CxC', 'type': 'ratio', 'category': 'Actividad'},
        {'row': 13, 'name': 'Rotación de inventarios', 'short': 'Rotación invent.', 'type': 'ratio', 'category': 'Actividad'},
    ]

    def bs_val(row, col_idx): return safe_float(ws_bs.cell(row=row, column=col_idx).value)
    def is_val(row, col_idx): return safe_float(ws_is.cell(row=row, column=col_idx).value)

    all_ratios = {r['row']: [] for r in ratios_info}

    for i in range(len(years)):
        col_idx = 3 + i
        bs = {k: bs_val(k, col_idx) for k in [20, 55, 13, 30, 69, 40, 78]}
        is_ = {k: is_val(k, col_idx) for k in [28, 4, 5]}

        for row, calc in [
            (4, bs[20] / bs[55] if bs[55] else 0),
            (5, (bs[20] - bs[13]) / bs[55] if bs[55] else 0),
            (6, bs[69] / bs[40] if bs[40] else 0),
            (7, bs[69] / bs[78] if bs[78] else 0),
            (8, is_[28] / is_[4] if is_[4] else 0),
            (9, is_[28] / bs[40] if bs[40] else 0),
            (10, is_[28] / bs[78] if bs[78] else 0),
            (11, is_[4] / bs[40] if bs[40] else 0)
        ]:
            all_ratios[row].append(calc)

        if i < len(years) - 1:
            prev_col = col_idx + 1
            cxc_avg = (sum(bs_val(r, col_idx) for r in list(range(8, 12)) + list(range(24, 28))) + 
                      sum(bs_val(r, prev_col) for r in list(range(8, 12)) + list(range(24, 28)))) / 2
            all_ratios[12].append(is_[4] / cxc_avg if cxc_avg else 0)
            
            inv_avg = ((bs[13] + bs_val(13, prev_col)) + (bs[30] + bs_val(30, prev_col))) / 2
            all_ratios[13].append((-is_[5]) / inv_avg if inv_avg else 0)
        else:
            all_ratios[12].append(0)
            all_ratios[13].append(0)

    # 🆕 GENERAR INFORME CON IA 
    if IA_DISPONIBLE:
        try:
            print("🤖 Generando análisis narrativo con Gemini...")
            informe_ia = generar_informe_ia(years, all_ratios)
            print(f"✅ Análisis generado exitosamente")
        except Exception as e:
            print(f"⚠️ Error al generar informe: {e}")
            informe_ia = None
    else:
        print("ℹ️ Generando PDF sin análisis de IA")
        informe_ia = None

    # 🔄 GENERAR PDF COMBINADO CON GRÁFICOS E INFORME (FORMATO HORIZONTAL)
    empresa_limpia = re.sub(r'[<>:"/\\|?*]', '', company_name)[:50]
    pdf_path = os.path.join(output_dir, f"ANALISIS_FINANCIERO_{empresa_limpia}_{min(years)}-{max(years)}.pdf")
    
    with PdfPages(pdf_path) as pdf:
        # 📄 PÁGINA DE PORTADA - FORMATO HORIZONTAL (11.69 x 8.27)
        fig = plt.figure(figsize=(11.69, 8.27))
        fig.patch.set_facecolor('white')
        
        fig.text(0.5, 0.65, "ANÁLISIS FINANCIERO INTEGRAL", ha='center', 
                fontsize=26, fontweight='bold', color='darkblue')
        fig.text(0.5, 0.58, company_name, ha='center', 
                fontsize=20, color='darkslategray', fontweight='medium')
        fig.text(0.5, 0.50, "Ratios Financieros y Análisis Narrativo", ha='center', 
                fontsize=16, color='gray')
        
        fig.text(0.5, 0.38, f"Período de Análisis: {min(years)} - {max(years)}", 
                ha='center', fontsize=15, fontweight='medium')
        
        fig.text(0.5, 0.20, "Generado por Analizador Financiero SMV", 
                ha='center', fontsize=12, color='gray')
        
        fig.add_artist(plt.Line2D([0.2, 0.8], [0.55, 0.55], color='darkblue', linewidth=2))
        fig.add_artist(plt.Line2D([0.2, 0.8], [0.35, 0.35], color='darkblue', linewidth=1))
        
        pdf.savefig(fig)
        plt.close(fig)
        
        # 📊 GRÁFICOS DE RATIOS - FORMATO HORIZONTAL (2 PÁGINAS)
        grupos_graficos = [
            {
                'title': 'Análisis de Liquidez y Endeudamiento', 
                'categories': ['Liquidez', 'Endeudamiento'], 
                'layout': (2, 2)
            },
            {
                'title': 'Análisis de Rentabilidad y Actividad', 
                'categories': ['Rentabilidad', 'Actividad'], 
                'layout': (2, 3)
            }
        ]
        
        for group in grupos_graficos:
            group_ratios = [r for r in ratios_info if r['category'] in group['categories']]
            n_rows, n_cols = group['layout']
            
            # FORMATO HORIZONTAL para gráficos
            fig, axes = plt.subplots(n_rows, n_cols, figsize=(11.69, 8.27), squeeze=False)
            fig.patch.set_facecolor('white')
            
            fig.suptitle(group['title'], fontsize=18, y=0.96, fontweight='bold', 
                        color='darkblue', ha='center')
            
            for i, r in enumerate(group_ratios):
                row, col = i // n_cols, i % n_cols
                ax = axes[row, col]
                
                # Ajustar valores para rotación (último año no tiene dato)
                if r['row'] in [12, 13]:
                    x_years = years[:-1]
                    values = all_ratios[r['row']][:-1]
                else:
                    x_years = years
                    values = all_ratios[r['row']]
                
                bars = ax.bar(np.arange(len(x_years)), values, 
                             width=0.65, color='#2E86AB', edgecolor='white', 
                             linewidth=1.2, alpha=0.9)
                
                ax.set_xticks(np.arange(len(x_years)))
                ax.set_xticklabels([str(y) for y in x_years], fontsize=11, fontweight='medium')
                
                ax.set_title(r['name'], fontsize=13, fontweight='bold', 
                           color='darkslategray', pad=12)
                
                nice_ticks(ax, values, r['type'])
                ax.set_ylabel(r['short'], fontsize=10, color='gray')
                
                ax.grid(axis='y', linestyle='--', alpha=0.3, color='gray')
                ax.set_axisbelow(True)
                
                add_value_labels(ax, bars, values, r['type'])
                
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.spines['left'].set_color('lightgray')
                ax.spines['bottom'].set_color('lightgray')
            
            # Remover subplots vacíos
            for i in range(len(group_ratios), n_rows * n_cols):
                row, col = i // n_cols, i % n_cols
                fig.delaxes(axes[row, col])
            
            plt.tight_layout(rect=[0.03, 0.03, 0.97, 0.92], 
                           pad=1.8, h_pad=2.5, w_pad=1.5)
            
            pdf.savefig(fig, dpi=300)
            plt.close(fig)
        
        # 📝 INFORME NARRATIVO - FORMATO VERTICAL (8.27 x 11.69)
        if informe_ia:
            import textwrap
            
            # Parámetros optimizados para A4 vertical
            caracteres_por_linea = 85
            lineas_por_pagina = 35
            y_inicial = 0.92
            altura_linea = 0.024
            margen_izquierdo = 0.08
            
            fig_info = plt.figure(figsize=(8.27, 11.69))
            fig_info.patch.set_facecolor('white')
            
            fig_info.text(0.5, 0.96, "ANÁLISIS FINANCIERO NARRATIVO", 
                         ha='center', fontsize=16, fontweight='bold', color='darkblue')
            fig_info.add_artist(plt.Line2D([0.1, 0.9], [0.94, 0.94], color='darkblue', linewidth=1))
            
            y_position = y_inicial
            lineas_en_pagina = 0
            
            lineas = informe_ia.split('\n')
            
            for linea in lineas:
                if not linea.strip():
                    y_position -= altura_linea * 0.5
                    lineas_en_pagina += 0.5
                    continue
                
                if len(linea) > caracteres_por_linea:
                    sub_lineas = textwrap.wrap(linea, caracteres_por_linea, break_long_words=False)
                else:
                    sub_lineas = [linea]
                
                for sub_linea in sub_lineas:
                    if lineas_en_pagina >= lineas_por_pagina:
                        pdf.savefig(fig_info)
                        plt.close(fig_info)
                        
                        fig_info = plt.figure(figsize=(8.27, 11.69))
                        fig_info.patch.set_facecolor('white')
                        y_position = 0.95
                        lineas_en_pagina = 0
                    
                    texto_limpio = sub_linea.strip()
                    
                    if (texto_limpio.endswith(':') or 
                        any(word in texto_limpio.upper() for word in ['ANÁLISIS', 'RESUMEN', 'RECOMENDACIONES', 'FORTALEZAS', 'ÁREAS'])):
                        fig_info.text(margen_izquierdo, y_position, texto_limpio, 
                                     ha='left', va='top', fontsize=11, fontweight='bold', color='darkblue')
                        lineas_en_pagina += 1.2 
                        
                    elif texto_limpio.startswith('•') or texto_limpio.startswith('-'):
                        fig_info.text(margen_izquierdo + 0.03, y_position, texto_limpio, 
                                     ha='left', va='top', fontsize=10, color='black')
                        lineas_en_pagina += 1
                        
                    else:
                        fig_info.text(margen_izquierdo, y_position, texto_limpio, 
                                     ha='left', va='top', fontsize=10, color='black')
                        lineas_en_pagina += 1
                    
                    y_position -= altura_linea
            
            pdf.savefig(fig_info)
            plt.close(fig_info)

    wb.close()
    print(f"✅ PDF generado: {pdf_path}")
    print(f"   - 1 página de portada (horizontal)")
    print(f"   - 2 páginas de gráficos de ratios (horizontal)")
    print(f"   - Páginas de análisis narrativo (vertical)")
    return pdf_path

def generate_complete_financial_pdf(report_path, output_dir="."):
    """Alias para generate_ratios_charts_pdf"""
    return generate_ratios_charts_pdf(report_path, output_dir)
